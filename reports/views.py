import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from datetime import date, timedelta
from admissions.models import Admission
from enquiries.models import Enquiry
from fees.models import Payment
from students.models import Student
from accounts.models import User
from accounts.decorators import role_required
from universities.models import University
from courses.models import Course

@login_required
@role_required('admin', 'accountant')
def reports_dashboard(request):
    return render(request, 'reports/dashboard.html')

@login_required
@role_required('admin', 'accountant')
def admission_reports(request):
    qs = Admission.objects.select_related('student', 'university', 'course', 'counsellor')

    period = request.GET.get('period', '').strip()
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    student_search = request.GET.get('student', '').strip()
    university_id = request.GET.get('university', '').strip()
    course_id = request.GET.get('course', '').strip()
    year = request.GET.get('year', '').strip()

    today = date.today()

    if period == 'today':
        qs = qs.filter(admission_date=today)
    elif period == 'yesterday':
        qs = qs.filter(admission_date=today - timedelta(days=1))
    elif period == 'week':
        week_start = today - timedelta(days=today.weekday())
        qs = qs.filter(admission_date__gte=week_start, admission_date__lte=today)
    elif period == 'month':
        qs = qs.filter(admission_date__gte=today.replace(day=1), admission_date__lte=today)
    elif period == 'last_month':
        first_this_month = today.replace(day=1)
        last_month_end = first_this_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        qs = qs.filter(admission_date__gte=last_month_start, admission_date__lte=last_month_end)
    elif period == 'year':
        qs = qs.filter(admission_date__gte=today.replace(month=1, day=1), admission_date__lte=today)
    elif period == 'custom' and from_date and to_date:
        from datetime import datetime as dt
        qs = qs.filter(admission_date__gte=dt.strptime(from_date, '%Y-%m-%d').date(), admission_date__lte=dt.strptime(to_date, '%Y-%m-%d').date())

    if student_search:
        qs = qs.filter(Q(student__name__icontains=student_search) | Q(student__student_id__icontains=student_search))
    if university_id:
        qs = qs.filter(university_id=university_id)
    if course_id:
        qs = qs.filter(course_id=course_id)
    if year:
        qs = qs.filter(admission_date__year=year)

    total_admissions = qs.count()
    total_students = qs.values('student').distinct().count()
    total_fees = qs.aggregate(t=Sum('total_fee'))['t'] or 0
    total_collected = Payment.objects.filter(is_voided=False, admission__in=qs).aggregate(t=Sum('amount'))['t'] or 0
    total_pending = total_fees - total_collected

    uni_stats = qs.values('university__id', 'university__name').annotate(
        count=Count('id'),
        fees=Sum('total_fee'),
    ).order_by('-count')

    uni_data = []
    for us in uni_stats:
        uni_collected = Payment.objects.filter(is_voided=False, admission__university_id=us['university__id'], admission__in=qs).aggregate(t=Sum('amount'))['t'] or 0
        uni_data.append({
            'id': us['university__id'],
            'name': us['university__name'],
            'count': us['count'],
            'fees': us['fees'] or 0,
            'collected': uni_collected,
            'pending': (us['fees'] or 0) - uni_collected,
        })

    import calendar
    monthly_data = []
    monthly_labels = []
    year_int = int(year) if year else today.year
    for m in range(1, 13):
        count = qs.filter(admission_date__year=year_int, admission_date__month=m).count()
        monthly_data.append(count)
        monthly_labels.append(calendar.month_abbr[m])

    years = qs.dates('admission_date', 'year').order_by('admission_date')
    yearly_data = []
    yearly_labels = []
    for y in years:
        yearly_labels.append(y.year)
        yearly_data.append(qs.filter(admission_date__year=y.year).count())

    universities = University.objects.filter(is_active=True).order_by('name')
    courses = Course.objects.filter(is_active=True).order_by('name')

    context = {
        'admissions': qs[:200],
        'total_admissions': total_admissions,
        'total_students': total_students,
        'total_universities': uni_stats.count(),
        'total_courses': qs.values('course').distinct().count(),
        'total_fees': total_fees,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'uni_data': uni_data,
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'yearly_labels': json.dumps(yearly_labels),
        'yearly_data': json.dumps(yearly_data),
        'universities': universities,
        'courses': courses,
        'period': period,
        'from_date': from_date,
        'to_date': to_date,
        'student_search': student_search,
        'selected_university': university_id,
        'selected_course': course_id,
        'selected_year': year,
        'current_year': today.year,
    }
    return render(request, 'reports/admissions.html', context)


@login_required
@role_required('admin')
def university_admission_detail(request, university_id):
    university = get_object_or_404(University, pk=university_id)
    admissions = Admission.objects.filter(university=university).select_related('student', 'course', 'counsellor')

    period = request.GET.get('period', '').strip()
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    today = date.today()

    if period == 'today':
        admissions = admissions.filter(admission_date=today)
    elif period == 'yesterday':
        admissions = admissions.filter(admission_date=today - timedelta(days=1))
    elif period == 'week':
        week_start = today - timedelta(days=today.weekday())
        admissions = admissions.filter(admission_date__gte=week_start, admission_date__lte=today)
    elif period == 'month':
        admissions = admissions.filter(admission_date__gte=today.replace(day=1), admission_date__lte=today)
    elif period == 'custom' and from_date and to_date:
        from datetime import datetime as dt
        admissions = admissions.filter(admission_date__gte=dt.strptime(from_date, '%Y-%m-%d').date(), admission_date__lte=dt.strptime(to_date, '%Y-%m-%d').date())

    total_students = admissions.values('student').distinct().count()

    courses = Course.objects.filter(university=university, is_active=True)
    course_data = []
    for course in courses:
        course_admissions = admissions.filter(course=course)
        count = course_admissions.count()
        if count > 0:
            fees = course_admissions.aggregate(t=Sum('total_fee'))['t'] or 0
            collected = Payment.objects.filter(is_voided=False, admission__in=course_admissions).aggregate(t=Sum('amount'))['t'] or 0
            course_data.append({
                'course': course,
                'count': count,
                'fees': fees,
                'collected': collected,
                'pending': fees - collected,
            })

    return render(request, 'reports/university_admission_detail.html', {
        'university': university,
        'admissions': admissions[:200],
        'total_students': total_students,
        'total_admissions': admissions.count(),
        'course_data': course_data,
        'period': period,
        'from_date': from_date,
        'to_date': to_date,
    })


@login_required
@role_required('admin', 'accountant')
def enquiry_reports(request):
    qs = Enquiry.objects.all()
    source = request.GET.get('source')
    status = request.GET.get('status')
    counsellor_id = request.GET.get('counsellor')
    if source:
        qs = qs.filter(source=source)
    if status:
        qs = qs.filter(status=status)
    if counsellor_id:
        qs = qs.filter(assigned_to_id=counsellor_id)
    total = qs.count()
    new_count = qs.filter(status='new').count()
    converted = qs.filter(status='converted').count()
    lost = qs.filter(status='lost').count()
    rate = round(converted / total * 100, 1) if total else 0
    counsellors = User.objects.filter(role__in=['counsellor', 'admin'])
    return render(request, 'reports/enquiries.html', {
        'enquiries': qs, 'total': total, 'new_count': new_count,
        'converted': converted, 'lost': lost, 'rate': rate, 'counsellors': counsellors,
    })

@login_required
@role_required('admin', 'accountant')
def fee_reports(request):
    payments = Payment.objects.filter(is_voided=False).select_related('admission__student', 'admission__university', 'admission__course', 'received_by')
    total_collected = payments.aggregate(total=Sum('amount'))['total'] or 0
    total_pending = Admission.objects.filter(status__in=['active', 'fee_pending']).aggregate(
        t=Sum('total_fee'))['t'] or 0
    total_paid = Payment.objects.filter(is_voided=False).aggregate(t=Sum('amount'))['t'] or 0
    pending = total_pending - total_paid
    return render(request, 'reports/fees.html', {
        'payments': payments, 'total_collected': total_collected,
        'total_pending': pending,
    })

@login_required
@role_required('admin', 'accountant')
def payment_reports(request):
    payments = Payment.objects.filter(is_voided=False).select_related(
        'admission__student', 'admission__university', 'admission__course', 'received_by'
    )

    period = request.GET.get('period', '').strip()
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    student_search = request.GET.get('student', '').strip()
    university_id = request.GET.get('university', '').strip()
    payment_mode = request.GET.get('payment_mode', '').strip()

    today = date.today()

    if period == 'today':
        payments = payments.filter(payment_date=today)
    elif period == 'yesterday':
        payments = payments.filter(payment_date=today - timedelta(days=1))
    elif period == 'week':
        week_start = today - timedelta(days=today.weekday())
        payments = payments.filter(payment_date__gte=week_start, payment_date__lte=today)
    elif period == 'month':
        payments = payments.filter(payment_date__gte=today.replace(day=1), payment_date__lte=today)
    elif period == 'last_month':
        first_this_month = today.replace(day=1)
        last_month_end = first_this_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        payments = payments.filter(payment_date__gte=last_month_start, payment_date__lte=last_month_end)
    elif period == 'year':
        payments = payments.filter(payment_date__gte=today.replace(month=1, day=1), payment_date__lte=today)
    elif period == 'custom' and from_date and to_date:
        from datetime import datetime as dt
        payments = payments.filter(payment_date__gte=dt.strptime(from_date, '%Y-%m-%d').date(), payment_date__lte=dt.strptime(to_date, '%Y-%m-%d').date())

    if student_search:
        payments = payments.filter(Q(admission__student__name__icontains=student_search) | Q(admission__student__student_id__icontains=student_search))
    if university_id:
        payments = payments.filter(admission__university_id=university_id)
    if payment_mode:
        payments = payments.filter(payment_mode=payment_mode)

    total_collected = payments.aggregate(t=Sum('amount'))['t'] or 0
    total_fees_all = Admission.objects.aggregate(t=Sum('total_fee'))['t'] or 0
    total_paid_all = Payment.objects.filter(is_voided=False).aggregate(t=Sum('amount'))['t'] or 0
    total_pending = total_fees_all - total_paid_all

    all_admissions = Admission.objects.all()
    total_due = 0
    for adm in all_admissions:
        bal = adm.balance_amount
        if bal > 0:
            total_due += bal

    mode_stats = payments.values('payment_mode').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-total')

    mode_display = dict(Payment.MODE_CHOICES)
    mode_data = []
    for ms in mode_stats:
        mode_data.append({
            'mode': ms['payment_mode'],
            'display': mode_display.get(ms['payment_mode'], ms['payment_mode']),
            'count': ms['count'],
            'total': ms['total'] or 0,
        })

    pending_admissions = Admission.objects.filter(
        status__in=['active', 'fee_pending']
    ).select_related('student', 'university', 'course')

    due_data = []
    for adm in pending_admissions:
        balance = adm.balance_amount
        if balance > 0:
            last_payment = Payment.objects.filter(admission=adm, is_voided=False).order_by('-payment_date').first()
            if last_payment:
                due_date = last_payment.payment_date + timedelta(days=30)
            else:
                due_date = adm.admission_date + timedelta(days=30)

            if today > due_date and balance > 0:
                status = 'overdue'
            elif balance <= 0:
                status = 'paid'
            elif adm.total_fee and balance < adm.total_fee:
                status = 'partial'
            else:
                status = 'pending'

            due_data.append({
                'admission': adm,
                'total_fee': adm.total_fee,
                'paid': adm.paid_amount,
                'pending': balance,
                'due_date': due_date,
                'status': status,
            })

    import calendar
    monthly_labels = []
    monthly_data = []
    for i in range(5, -1, -1):
        d = today - timedelta(days=i*30)
        month_start = d.replace(day=1)
        if d.month == 12:
            month_end = d.replace(year=d.year+1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = d.replace(month=d.month+1, day=1) - timedelta(days=1)
        amount = Payment.objects.filter(is_voided=False, payment_date__gte=month_start, payment_date__lte=month_end).aggregate(t=Sum('amount'))['t'] or 0
        monthly_data.append(float(amount))
        monthly_labels.append(calendar.month_abbr[d.month])

    universities = University.objects.filter(is_active=True).order_by('name')

    context = {
        'payments': payments[:200],
        'total_collected': total_collected,
        'total_fees': total_fees_all,
        'total_pending': total_pending,
        'total_due': total_due,
        'mode_data': mode_data,
        'due_data': due_data[:100],
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'universities': universities,
        'payment_modes': Payment.MODE_CHOICES,
        'period': period,
        'from_date': from_date,
        'to_date': to_date,
        'student_search': student_search,
        'selected_university': university_id,
        'selected_mode': payment_mode,
    }
    return render(request, 'reports/payments.html', context)


@login_required
@role_required('admin')
def import_students(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Please select an Excel file.')
            return redirect('import_students')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Only .xlsx or .xls files are supported.')
            return redirect('import_students')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            headers_lower = [str(h).strip().lower() if h else '' for h in headers]

            name_idx = None
            mobile_idx = None
            for i, h in enumerate(headers_lower):
                if h in ('name', 'student name', 'student_name', 'studentname'):
                    name_idx = i
                elif h in ('mobile', 'phone', 'mobile number', 'mobile_number', 'phonenumber', 'contact'):
                    mobile_idx = i

            if name_idx is None or mobile_idx is None:
                messages.error(request, 'Excel must have columns: "Name" (or "Student Name") and "Mobile" (or "Phone").')
                return redirect('import_students')

            created = 0
            skipped = 0
            errors = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = str(row[name_idx] or '').strip()
                mobile = str(row[mobile_idx] or '').strip()

                if not name or not mobile:
                    skipped += 1
                    continue

                mobile = mobile.replace(' ', '').replace('-', '')
                if len(mobile) == 12 and mobile.startswith('91'):
                    mobile = mobile[2:]
                if len(mobile) != 10 or mobile[0] not in '6789':
                    errors.append(f'Row {row_num}: Invalid mobile "{mobile}"')
                    skipped += 1
                    continue

                if Student.objects.filter(mobile=mobile).exists():
                    skipped += 1
                    continue

                Student.objects.create(name=name, mobile=mobile, status='prospect')
                created += 1

            wb.close()

            if errors:
                messages.warning(request, f'Imported {created} students. Skipped {skipped}. Errors: {" | ".join(errors[:5])}')
            else:
                messages.success(request, f'Successfully imported {created} students. Skipped {skipped} (duplicates/empty).')
        except Exception as e:
            messages.error(request, f'Error reading file: {str(e)}')

        return redirect('import_students')

    return render(request, 'reports/import_students.html')


@login_required
@role_required('admin', 'accountant')
def export_students_excel(request):
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    ws.append(['Name', 'Mobile'])

    students = Student.objects.order_by('student_id')
    for s in students:
        ws.append([
            s.name,
            s.mobile,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students_list.xlsx"'
    wb.save(response)
    wb.close()
    return response
